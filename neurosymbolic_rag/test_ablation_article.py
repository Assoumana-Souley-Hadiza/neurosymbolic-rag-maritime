"""
Script d'étude d'ablation du système RAG
Compare 4 configurations de pipeline pour mesurer l'impact de chaque composant.

Architecture confirmée (identique à batch_query.py) :
  - Le Graph Retriever N'EST PAS utilisé côté retrieval.
  - Le Neo4j sert uniquement à :
      (a) l'expansion de requête  → expanded_terms / synonym_sets  (dense + sparse)
      (b) l'enrichissement LLM    → ontology_agent.enrich()

Les 4 pipelines isolent donc :

  ┌─────────────────┬────────┬────────┬────────────────┬────────────────────┐
  │ Pipeline        │ Dense  │ Sparse │ Expansion Neo4j│ Enrichiss. Ontolog.│
  ├─────────────────┼────────┼────────┼────────────────┼────────────────────┤
  │ Classique       │ brut   │  ✗     │      ✗         │        ✗           │
  │ Sans_Sparse     │ brut   │  ✗     │      ✗         │        ✓           │
  │ Sans_Ontologie  │ brut   │  ✓     │      ✗         │        ✗           │
  │ Complet         │ enrichi│  ✓     │      ✓         │        ✓           │
  └─────────────────┴────────┴────────┴────────────────┴────────────────────┘

Comparaisons utiles :
  Classique     → Sans_Sparse    : apport de l'expansion Neo4j + ontologie seuls
  Sans_Sparse   → Complet        : apport du sparse retriever
  Sans_Ontologie→ Complet        : apport de l'enrichissement ontologique
  Classique     → Complet        : gain global du pipeline neuro

Sortie Excel : UN SEUL fichier avec plusieurs feuilles
  - "Synthèse"             : taux OUI/NON global par pays et par pipeline
  - Une feuille par interdiction (toutes les questions × tous les pays)
  - Une feuille par pays   : toutes les interdictions pour ce pays
"""

import sys
import csv
import json
import logging
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import argparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from rag.core.fusion import HybridFusion
from rag.core.query_analyzer import QueryAnalyzer
from rag.llm_generator import LLMGenerator
from rag.integration.neo4j_bridge import Neo4jBridge
from rag.core.retrievers import DenseRetriever, SparseRetriever
from rag.neo4j_ontology_agent import Neo4jOntologyAgent

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('batch_ablation.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# PAYS
# ─────────────────────────────────────────────────────────────────────────────

PAYS = {
    "Bénin":         {"code": "ben", "prep": "au Bénin",          "prep_q": "au Bénin"},
    "Cameroun":      {"code": "cmr", "prep": "au Cameroun",       "prep_q": "au Cameroun"},
    "Togo":          {"code": "tog", "prep": "au Togo",           "prep_q": "au Togo"},
    "Maroc":         {"code": "mor", "prep": "au Maroc",          "prep_q": "au Maroc"},
    "Sénégal":       {"code": "sen", "prep": "au Sénégal",        "prep_q": "au Sénégal"},
    "Congo":         {"code": "cng", "prep": "au Congo",          "prep_q": "au Congo"},
    "Madagascar":    {"code": "mad", "prep": "à Madagascar",      "prep_q": "à Madagascar"},
}

# ─────────────────────────────────────────────────────────────────────────────
# QUESTIONS  (toutes les interdictions du document source)
# ─────────────────────────────────────────────────────────────────────────────

INTERDICTIONS = {
    "Chasse à la baleine": [
        "Est-ce qu'il existe un article portant sur l'interdiction de la chasse des baleines {prep} ?",
        "Est-ce qu'il existe un article portant sur l'interdiction de la chasse des baleines qui précise que cette mesure n'est applicable qu'à certaines zones, aires ou régions {prep} ?",
        "Est-ce qu'il existe un article portant sur l'interdiction de la chasse des baleines qui précise que cette mesure n'est pas applicable en permanence (temporalité) {prep} ?",
        "Est-ce qu'un texte juridique mentionne des exceptions à l'interdiction de la chasse des baleines {prep} ? Vérifie si ce texte inclut des exceptions relatives à d'autres domaines que de la santé, de l'ordre public et de la recherche.",
        "Est-ce qu'un texte juridique mentionne des exceptions à l'interdiction de la chasse des baleines {prep} ? Ne pas inclure les délais de mise en place de la loi. Les dérogations ou autorisations spécifiques constituent des exceptions.",
        "Est-ce qu'il existe un article qui précise que l'infraction à l'interdiction de la chasse des baleines {prep} entraîne une sanction financière ?",
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction de la chasse des baleines ?"
    ],
    "Construction sur le littoral": [
        "Est-ce qu'il existe un article portant sur l'interdiction de la construction sur le littoral {prep} ?",
        "Est-ce que le texte juridique mentionne des exceptions à l'interdiction de la construction sur le littoral {prep} ?"
    ],
    "Rejet d'hydrocarbures": [
        "Est-ce qu'il existe un article portant sur l'interdiction du rejet d'hydrocarbures {prep} ?",
        "Est-ce que le rejet d'hydrocarbures {prep} entraîne une peine de prison ?",
        "Est-ce que le rejet d'hydrocarbures {prep} entraîne une sanction financière (amende) ?"
    ],
     "Chalutage de fond": [
        "Est-ce qu'il existe un article portant sur l'interdiction du chalutage de fond {prep} ?",
        "Est-ce qu'un texte juridique mentionne des exceptions à l'interdiction du chalutage de fond {prep} ?"
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction du chalutage de fond ?",
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction du chalutage de fond ? Si ces procédures existent, vérifiez si elles précisent des périodes (temporalité) spécifiques pour le contrôle du respect de l'interdiction du chalutage de fond.",
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction du chalutage de fond ? Si ces procédures existent, vérifiez si elles précisent des lieux ou zones spécifiques pour le contrôle du respect de l'interdiction du chalutage de fond.",
    ],
    "Oiseaux Marins": [
        "Est-ce qu'il existe un article portant sur l'interdiction des oiseaux marins {prep} ?",
        "Est-ce qu'il existe un article portant sur l'interdiction des oiseaux marins qui précise que cette mesure n'est applicable qu'à certaines zones, aires ou régions {prep} ?",
        "Est-ce qu'il existe un article portant sur l'interdiction des oiseaux marins qui précise que cette mesure n'est pas applicable en permanence (temporalité) {prep} ?",
        "Est-ce qu'un texte juridique précise les types d'activités concernées par l'interdiction des oiseaux marins {prep} ?",
        "Est-ce qu'un texte juridique mentionne des exceptions à l'interdiction des oiseaux marins {prep} ? Vérifie si ce texte inclut des exceptions relatives à d'autres domaines que ceux de la santé, de l'ordre public et de la recherche.",
        "Est-ce qu'un texte juridique mentionne des exceptions à l'interdiction des oiseaux marins {prep} ?",
    ],
    "Extraction de sable marin": [
        "Est-ce qu'il existe un article portant sur l'interdiction d'extraction de sable marin {prep} ?",
        "Est-ce qu'il existe un article portant sur l'interdiction d'extraction de sable marin {prep} qui précise que cette mesure n'est pas applicable en permanence (temporalité) ?",
        "Est-ce que l'extraction de sable sur le littoral {prep} entraîne une sanction financière  ?",
        "Est-ce que l'extraction de sable sur le littoral {prep} entraîne une peine de prison ?",
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction de l'extraction de sable sur le littoral ?",
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction de l'extraction de sable sur le littoral ? Si ces procédures existent, vérifiez si elles précisent des périodes (temporalité) spécifiques pour le contrôle du respect de l'interdiction de l'extraction de sable sur le littoral.",
    ]
}


# ─────────────────────────────────────────────────────────────────────────────
# DÉFINITION DES PIPELINES D'ABLATION
# ─────────────────────────────────────────────────────────────────────────────

ABLATION_PIPELINES: List[Dict] = [
    {
        "name":          "Classique",
        "label":         "RAG Classique\n(dense pur)",
        "use_sparse":    False,
        "use_expansion": False,
        "use_ontology":  False,
        "header_color":  "2F4F8F",
    },
    {
        "name":          "Complet",
        "label":         "RAG Complet\n(dense enrichi + sparse + ontologie)",
        "use_sparse":    True,
        "use_expansion": True,
        "use_ontology":  True,
        "header_color":  "4A7C59",
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# UTILITAIRES EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def _slugify(text: str) -> str:
    normalized = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('utf-8')
    return normalized.lower().replace(" ", "_").replace("'", "")


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Retire les caractères interdits dans les noms de feuilles Excel."""
    forbidden = r"\/*?:[]"
    for ch in forbidden:
        name = name.replace(ch, "_")
    return name[:max_len]


def _clean_for_excel(val) -> str:
    if not isinstance(val, str):
        return val
    val = re.sub(r'[\x01-\x08]', '• ', val)
    return re.sub(r'[\x00\x0b\x0c\x0e-\x1f]', '', val)


_HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN   = Alignment(vertical="top", wrap_text=True)
_BORDER_THIN  = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
_ALT_FILL = PatternFill("solid", start_color="F2F2F2")


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color)


def _style_sheet(ws, col_widths: List[int], header_fills: List[PatternFill]):
    """Applique styles, largeurs, freeze et bandes de couleur alternées."""
    for col_idx, (cell, fill) in enumerate(zip(ws[1], header_fills), 1):
        cell.font      = _HEADER_FONT
        cell.fill      = fill
        cell.alignment = _HEADER_ALIGN
        cell.border    = _BORDER_THIN

    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for cell in row:
            cell.alignment = _CELL_ALIGN
            cell.border    = _BORDER_THIN
            if fill:
                cell.fill = fill

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 80


# ─────────────────────────────────────────────────────────────────────────────
# GESTION DU FICHIER EXCEL UNIQUE
# ─────────────────────────────────────────────────────────────────────────────
#
# Structure du fichier (feuilles dans l'ordre) :
#   0. "Synthèse"                  — taux OUI/NON global par pays × pipeline
#   1. "Chalutage de fond"         — toutes questions × tous pays pour cette interdiction
#   2. "Chasse à la baleine"       — idem
#   3. "Construction côtière"      — idem
#   4. "Extraction de sable"       — idem
#   5. "Oiseaux marins"            — idem
#   6. "Rejet d'hydrocarbures"     — idem
#   7. "Bénin"                     — toutes interdictions pour ce pays
#   8. "Cameroun"                  — idem
#   …  (une feuille par pays)
#
# Les feuilles par interdiction regroupent les données de TOUS les pays ;
# les feuilles par pays regroupent TOUTES les interdictions.
#
# En-têtes des feuilles de données :
#   Pays | Interdiction | N° | Question | [Pipeline 1] | … | [Pipeline 4]
#
# En-têtes de la feuille Synthèse :
#   Pays | Total | OUI Classique | NON Classique | … | OUI Complet | NON Complet

def _build_response_headers() -> Tuple[List[str], List[PatternFill]]:
    """Retourne les en-têtes et les fills colorés pour une feuille de données."""
    headers = ["Pays", "Interdiction", "N°", "Question"] + [p["label"] for p in ABLATION_PIPELINES]
    fills = (
        [_make_fill("1F4E79")] * 4          # colonnes d'identification
        + [_make_fill(p["header_color"]) for p in ABLATION_PIPELINES]
    )
    return headers, fills


def _data_col_widths() -> List[int]:
    n = len(ABLATION_PIPELINES)
    response_w = max(10, 220 // n)
    return [18, 22, 5, 55] + [response_w] * n


def _row_to_cells(r: Dict) -> List:
    cells = [
        _clean_for_excel(r.get("Pays", "")),
        _clean_for_excel(r.get("Interdiction", "")),
        r.get("Numero_Question", ""),
        _clean_for_excel(r.get("Question", "")),
    ]
    for p in ABLATION_PIPELINES:
        cells.append(_clean_for_excel(r.get(f"Reponse_{p['name']}", "N/A")))
    return cells


def _write_data_sheet(ws, rows: List[Dict]):
    headers, fills = _build_response_headers()
    ws.append(headers)
    ws.row_dimensions[1].height = 40
    for r in rows:
        ws.append(_row_to_cells(r))
    _style_sheet(ws, col_widths=_data_col_widths(), header_fills=fills)


def _write_summary_sheet(ws, all_rows: List[Dict]):
    """Feuille Synthèse : OUI / NON par pays et par pipeline."""
    headers = ["Pays", "Total"]
    for p in ABLATION_PIPELINES:
        headers += [f"OUI — {p['name']}", f"NON — {p['name']}"]
    ws.append(headers)

    for pays_nom in PAYS.keys():
        pays_rows = [r for r in all_rows if r["Pays"] == pays_nom]
        if not pays_rows:
            continue
        row_data: List = [pays_nom, len(pays_rows)]
        for p in ABLATION_PIPELINES:
            col = f"Reponse_{p['name']}"
            oui = sum(1 for r in pays_rows if "OUI" in str(r.get(col, "")).upper())
            non = sum(1 for r in pays_rows if "NON" in str(r.get(col, "")).upper())
            row_data += [oui, non]
        ws.append(row_data)

    # Ligne de totaux
    total_row: List = ["TOTAL", len(all_rows)]
    for p in ABLATION_PIPELINES:
        col = f"Reponse_{p['name']}"
        total_row += [
            sum(1 for r in all_rows if "OUI" in str(r.get(col, "")).upper()),
            sum(1 for r in all_rows if "NON" in str(r.get(col, "")).upper()),
        ]
    ws.append(total_row)

    # Style
    summary_fills = [_make_fill("1F4E79"), _make_fill("1F4E79")]
    for p in ABLATION_PIPELINES:
        summary_fills += [_make_fill(p["header_color"]), _make_fill(p["header_color"])]
    col_widths_s = [20, 8] + [18] * (2 * len(ABLATION_PIPELINES))
    _style_sheet(ws, col_widths=col_widths_s, header_fills=summary_fills)


def ecrire_excel_unique(output_path: Path, resultats: List[Dict]):
    """
    Crée (ou recrée) le fichier Excel unique avec :
      - une feuille "Synthèse" en première position
      - une feuille par interdiction (toutes avec tous les pays)
      - une feuille par pays (toutes les interdictions)
    """
    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    # ── 1. Feuille Synthèse ───────────────────────────────────────────────
    ws_synth = wb.create_sheet(title="Synthèse")
    _write_summary_sheet(ws_synth, resultats)

    # ── 2. Une feuille par interdiction ───────────────────────────────────
    for interdiction in INTERDICTIONS.keys():
        rows = [r for r in resultats if r["Interdiction"] == interdiction]
        if not rows:
            continue
        # Tri stable : pays dans l'ordre de PAYS, puis numéro de question
        rows_sorted = sorted(
            rows,
            key=lambda r: (list(PAYS.keys()).index(r["Pays"]), r["Numero_Question"]),
        )
        sheet_name = _safe_sheet_name(interdiction)
        ws = wb.create_sheet(title=sheet_name)
        _write_data_sheet(ws, rows_sorted)

    # ── 3. Une feuille par pays ───────────────────────────────────────────
    for pays_nom in PAYS.keys():
        rows = [r for r in resultats if r["Pays"] == pays_nom]
        if not rows:
            continue
        # Tri stable : interdiction dans l'ordre de INTERDICTIONS, puis numéro de question
        inter_order = list(INTERDICTIONS.keys())
        rows_sorted = sorted(
            rows,
            key=lambda r: (
                inter_order.index(r["Interdiction"]) if r["Interdiction"] in inter_order else 99,
                r["Numero_Question"],
            ),
        )
        sheet_name = _safe_sheet_name(pays_nom)
        ws = wb.create_sheet(title=sheet_name)
        _write_data_sheet(ws, rows_sorted)

    wb.save(output_path)
    logger.info(f"✓ Excel unique sauvegardé : {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# CLASSE PRINCIPALE
# ─────────────────────────────────────────────────────────────────────────────

class AblationQuerySystem:
    """
    Orchestre l'étude d'ablation en exécutant chaque pipeline RAG séparément.
    Toutes les sorties sont consolidées dans UN SEUL fichier Excel multi-feuilles.
    """

    # Les mots de bruit et keywords anglais sont gérés par l'OntologyAgent.
    def __init__(self):
        self.output_dir = Path("results/ablation_study")
        self.output_dir.mkdir(parents=True, exist_ok=True)

        existing_runs = []
        for f in self.output_dir.glob("ablation_run*.xlsx"):
            try:
                existing_runs.append(int(f.stem.split("_run")[-1].split("_")[0]))
            except ValueError:
                pass
        self.run_id = max(existing_runs) + 1 if existing_runs else 1

        self.run_timestamp  = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.questions_file = self.output_dir / "questions_generees.csv"
        self.results_csv    = self.output_dir / f"ablation_resultats_run{self.run_id}_{self.run_timestamp}.csv"
        self.summary_file   = self.output_dir / f"ablation_resume_run{self.run_id}_{self.run_timestamp}.json"

        # Fichier Excel UNIQUE pour ce run
        self.excel_file = self.output_dir / f"ablation_run{self.run_id}_{self.run_timestamp}.xlsx"

        # Composants RAG
        self.fusion           = None
        self.llm_gen          = None
        self.query_analyzer   = None
        self.neo4j_bridge     = None
        self.ontology_agent   = None
        self.dense_retriever  = None
        self.sparse_retriever = None

        logger.info(f"AblationQuerySystem — run #{self.run_id}")
        logger.info(f"Fichier Excel de sortie : {self.excel_file}")

    # ─────────────────────────────────────────────────────────────────────────
    # Initialisation
    # ─────────────────────────────────────────────────────────────────────────

    def init_rag_system(self) -> bool:
        try:
            logger.info("Initialisation des composants RAG...")

            self.dense_retriever  = DenseRetriever()
            self.sparse_retriever = SparseRetriever()

            self.neo4j_bridge = Neo4jBridge.from_config()
            if not self.neo4j_bridge.is_ready():
                logger.warning("⚠ Neo4j Bridge non connecté — expansion et ontologie ignorés.")
            else:
                logger.info("✓ Neo4j Bridge connecté")

            self.ontology_agent = Neo4jOntologyAgent(self.neo4j_bridge)
            self.fusion         = HybridFusion()
            self.query_analyzer = QueryAnalyzer()
            self.llm_gen        = LLMGenerator()

            if not self.llm_gen.is_ollama_available():
                logger.warning("⚠ LLM non disponible")
                return False

            logger.info("✓ Tous les composants RAG prêts")
            return True

        except Exception as e:
            logger.error(f"❌ Erreur d'initialisation : {e}", exc_info=True)
            return False

    # ─────────────────────────────────────────────────────────────────────────
    # Génération des questions
    # ─────────────────────────────────────────────────────────────────────────

    def generer_questions(self) -> List[Dict]:
        questions = []
        for pays_nom, pays_info in PAYS.items():
            prep, prep_q = pays_info["prep"], pays_info["prep_q"]
            for interdiction, templates in INTERDICTIONS.items():
                for num_q, template in enumerate(templates, 1):
                    questions.append({
                        "Pays":            pays_nom,
                        "Code_Pays":       pays_info["code"],
                        "Interdiction":    interdiction,
                        "Numero_Question": num_q,
                        "Question":        template.replace("{prep}", prep).replace("{prep_q}", prep_q),
                    })

        with open(self.questions_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=questions[0].keys())
            writer.writeheader()
            writer.writerows(questions)

        print(f"✓ {len(questions)} questions générées → {self.questions_file}")
        return questions

    # ─────────────────────────────────────────────────────────────────────────
    # Expansion Neo4j
    # ─────────────────────────────────────────────────────────────────────────

    def _build_expansion(self, question: str, intent: str = None) -> Tuple[Dict, List[str], set, Optional[Dict]]:
        synonym_sets:   Dict = {}
        expanded_terms: List = []
        graph_entities: set  = set()

        if not (self.ontology_agent and self.ontology_agent.is_ready()):
            return synonym_sets, expanded_terms, graph_entities, None

        query_context = self.ontology_agent.prepare_query(question, intent=intent)
        synonym_sets = query_context.get("synonyms_map", {})
        graph_entities = query_context.get("graph_entities", set())
        expanded_terms = query_context.get("expanded_terms", [])

        return synonym_sets, expanded_terms, graph_entities, query_context

    # ─────────────────────────────────────────────────────────────────────────
    # Exécution d'un pipeline
    # ─────────────────────────────────────────────────────────────────────────

    def _run_pipeline(
        self,
        pipeline:        Dict,
        question:        str,
        intent:          str,
        weights:         Dict,
        country_filter,
        category_filter,
        synonym_sets:    Dict,
        expanded_terms:  List[str],
        graph_entities:  set,
        query_context:   Optional[Dict] = None,
    ) -> str:
        _K = 12

        retriever_results: Dict = {}

        if self.dense_retriever and self.dense_retriever.is_ready():
            dense_kwargs = {"top_k": _K}
            if pipeline["use_expansion"] and expanded_terms:
                dense_kwargs["expanded_terms"] = expanded_terms
            retriever_results["dense"] = self.dense_retriever.retrieve(question, **dense_kwargs)

        if pipeline["use_sparse"] and self.sparse_retriever and self.sparse_retriever.is_ready():
            sparse_kwargs = {"top_k": _K}
            if pipeline["use_expansion"] and expanded_terms:
                sparse_kwargs["expanded_terms"] = expanded_terms
            retriever_results["sparse"] = self.sparse_retriever.retrieve(question, **sparse_kwargs)

        fusion_kwargs: Dict = {
            "retriever_results": retriever_results,
            "weights":           weights,
            "synonym_sets":      synonym_sets if pipeline["use_expansion"] else {},
            "top_k":             5,
            "country_filter":    country_filter,
            "query":             question,
            "disable_reranker":  pipeline["name"] == "Classique",
            "intent":            intent,
        }

        if pipeline["use_expansion"] and graph_entities:
            fusion_kwargs["graph_entities"] = graph_entities
            fusion_kwargs["ontology_boost"] = 0.5

        if pipeline["use_expansion"] and category_filter:
            fusion_kwargs["category_filter"] = category_filter

        fusion_results = self.fusion.fuse(**fusion_kwargs)

        enriched_context = None
        if pipeline["use_ontology"] and self.ontology_agent and self.ontology_agent.is_ready():
            try:
                enriched_context = self.ontology_agent.enrich(question, fusion_results, query_context=query_context)
            except Exception as e:
                logger.warning(f"[{pipeline['name']}] Ontology agent ignoré : {e}")

        response_gen = self.llm_gen.generate(
            query=question,
            fusion_results=fusion_results,
            intent=intent,
            enriched_context=enriched_context,
            stream=True,
        )

        if response_gen.get("error"):
            return "ERREUR"

        full_ans = "".join(list(response_gen["stream_generator"]))
        return re.sub(r'<analyse>.*?(?:</analyse>|$)', '', full_ans, flags=re.DOTALL).strip()

    # ─────────────────────────────────────────────────────────────────────────
    # Traitement d'une question
    # ─────────────────────────────────────────────────────────────────────────

    def interroger_question(self, question_data: Dict) -> Dict:
        question = question_data["Question"]

        base = {
            "Pays":            question_data["Pays"],
            "Code_Pays":       question_data["Code_Pays"],
            "Interdiction":    question_data["Interdiction"],
            "Numero_Question": question_data["Numero_Question"],
            "Question":        question,
            "Modele":          "N/A",
            "Timestamp":       datetime.now().isoformat(),
            "Erreur":          "",
        }
        for p in ABLATION_PIPELINES:
            base[f"Reponse_{p['name']}"] = "ERREUR"

        try:
            if not self.query_analyzer or not self.fusion or not self.llm_gen:
                raise RuntimeError("Système RAG non initialisé")

            base["Modele"] = self.llm_gen.model_name

            intent, weights, country_filter, category_filter = \
                self.query_analyzer.analyze(question)

            needs_expansion = any(p["use_expansion"] for p in ABLATION_PIPELINES)
            if needs_expansion:
                synonym_sets, expanded_terms, graph_entities, query_context = self._build_expansion(question, intent=intent)
            else:
                synonym_sets, expanded_terms, graph_entities, query_context = {}, [], set(), None

            errors = []
            for pipeline in ABLATION_PIPELINES:
                try:
                    answer = self._run_pipeline(
                        pipeline        = pipeline,
                        question        = question,
                        intent          = intent,
                        weights         = weights,
                        country_filter  = country_filter,
                        category_filter = category_filter,
                        synonym_sets    = synonym_sets,
                        expanded_terms  = expanded_terms,
                        graph_entities  = graph_entities,
                        query_context   = query_context,
                    )
                    base[f"Reponse_{pipeline['name']}"] = answer
                    logger.debug(f"[{pipeline['name']}] ✓")
                except Exception as e:
                    logger.error(f"[{pipeline['name']}] Erreur pipeline : {e}", exc_info=True)
                    errors.append(f"{pipeline['name']}: {e}")

            if errors:
                base["Erreur"] = " | ".join(errors)

        except Exception as e:
            logger.error(f"Erreur traitement question : {e}", exc_info=True)
            base["Erreur"] = str(e)

        return base

    # ─────────────────────────────────────────────────────────────────────────
    # Sauvegarde CSV
    # ─────────────────────────────────────────────────────────────────────────

    def sauvegarder_resultats_csv(self, resultats: List[Dict]):
        if not resultats:
            return
        fieldnames = [
            "Pays", "Code_Pays", "Interdiction", "Numero_Question", "Question",
            *[f"Reponse_{p['name']}" for p in ABLATION_PIPELINES],
            "Modele", "Erreur", "Timestamp",
        ]
        with open(self.results_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(resultats)

    # ─────────────────────────────────────────────────────────────────────────
    # Résumé JSON
    # ─────────────────────────────────────────────────────────────────────────

    def generer_resume(self, resultats: List[Dict]) -> Dict:
        def _stats(rows: List[Dict]) -> Dict:
            out: Dict = {
                "total":   len(rows),
                "erreurs": sum(
                    1 for r in rows
                    if any(r.get(f"Reponse_{p['name']}") == "ERREUR" for p in ABLATION_PIPELINES)
                ),
            }
            for p in ABLATION_PIPELINES:
                col = f"Reponse_{p['name']}"
                out[f"oui_{p['name']}"] = sum(1 for r in rows if "OUI" in str(r.get(col, "")).upper())
                out[f"non_{p['name']}"] = sum(1 for r in rows if "NON" in str(r.get(col, "")).upper())
            return out

        resume = {
            "date_execution": datetime.now().isoformat(),
            "run_id":         self.run_id,
            "excel_file":     str(self.excel_file),
            "pipelines":      [
                {"name": p["name"], "config": {
                    "use_sparse":    p["use_sparse"],
                    "use_expansion": p["use_expansion"],
                    "use_ontology":  p["use_ontology"],
                }}
                for p in ABLATION_PIPELINES
            ],
            **_stats(resultats),
            "par_pays":         {
                pays: _stats([r for r in resultats if r["Pays"] == pays])
                for pays in PAYS
            },
            "par_interdiction": {
                inter: _stats([r for r in resultats if r["Interdiction"] == inter])
                for inter in INTERDICTIONS
            },
        }

        with open(self.summary_file, 'w', encoding='utf-8') as f:
            json.dump(resume, f, ensure_ascii=False, indent=2)

        logger.info(f"✓ Résumé JSON → {self.summary_file}")
        return resume

    def afficher_resume(self, resume: Dict):
        print("\n" + "=" * 72)
        print("📊  RÉSUMÉ DE L'ÉTUDE D'ABLATION")
        print("=" * 72)
        print(f"Total questions : {resume['total']}   |   Erreurs : {resume['erreurs']}")
        print()

        col_w = 22
        descriptions = {
            "Classique":      "dense pur, sans expansion ni ontologie",
            "Sans_Sparse":    "dense enrichi + ontologie, sans sparse",
            "Sans_Ontologie": "dense enrichi + sparse, sans ontologie",
            "Complet":        "dense enrichi + sparse + ontologie",
        }

        print(f"{'Pipeline':<{col_w}} {'OUI':>6} {'NON':>6}   Description")
        print("-" * 72)
        for p in ABLATION_PIPELINES:
            n   = p["name"]
            oui = resume.get(f"oui_{n}", 0)
            non = resume.get(f"non_{n}", 0)
            print(f"{n:<{col_w}} {oui:>6} {non:>6}   {descriptions.get(n, '')}")

        print("=" * 72)
        print(f"\n📁  Excel unique → {resume.get('excel_file', '')}\n")

    # ─────────────────────────────────────────────────────────────────────────
    # Pipeline principal
    # ─────────────────────────────────────────────────────────────────────────

    def run(self, generate_only: bool = False, query_only: bool = False):
        try:
            if not query_only:
                questions = self.generer_questions()
            else:
                questions = []
                with open(self.questions_file, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        row["Numero_Question"] = int(row["Numero_Question"])
                        questions.append(row)
                print(f"✓ {len(questions)} questions chargées depuis {self.questions_file}")

            if generate_only:
                print("\n✅ Génération terminée. Utilisez --query-only pour lancer les pipelines.")
                return

            if not self.init_rag_system():
                logger.error("❌ Impossible d'initialiser le système RAG")
                return

            resultats: List[Dict] = []
            total   = len(questions)
            n_pipes = len(ABLATION_PIPELINES)

            print(f"\n🔬 Étude d'ablation — {total} questions × {n_pipes} pipelines")
            print(f"📁  Fichier Excel : {self.excel_file}\n")
            for p in ABLATION_PIPELINES:
                flag_s = "✓ sparse"    if p["use_sparse"]    else "✗ sparse"
                flag_e = "✓ expansion" if p["use_expansion"] else "✗ expansion"
                flag_o = "✓ ontologie" if p["use_ontology"]  else "✗ ontologie"
                print(f"  [{p['name']:<16}]  {flag_s}  {flag_e}  {flag_o}")
            print()

            for idx, q_data in enumerate(questions, 1):
                label = (
                    f"[{idx}/{total}] {q_data['Pays']} — "
                    f"{q_data['Interdiction']} (Q{q_data['Numero_Question']})"
                )
                print(label, end=" ", flush=True)

                result = self.interroger_question(q_data)
                resultats.append(result)
                print("✓")

                # Sauvegarde incrémentale après chaque question
                self.sauvegarder_resultats_csv(resultats)
                ecrire_excel_unique(self.excel_file, resultats)

                if idx % 10 == 0:
                    logger.info(f"Progression : {idx}/{total} questions traitées")

            resume = self.generer_resume(resultats)
            self.afficher_resume(resume)
            print(f"✅ Résultats dans : {self.output_dir}")

        except Exception as e:
            logger.error(f"❌ Erreur fatale : {e}", exc_info=True)
            raise


# ─────────────────────────────────────────────────────────────────────────────
# POINT D'ENTRÉE
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Étude d'ablation RAG — sortie dans un seul fichier Excel multi-feuilles",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="\n".join(
            f"  {p['name']:<18} {p['label'].replace(chr(10), ' ')}"
            for p in ABLATION_PIPELINES
        ),
    )
    parser.add_argument(
        "--generate-only", action="store_true",
        help="Générer les questions uniquement (pas d'interrogation)",
    )
    parser.add_argument(
        "--query-only", action="store_true",
        help="Utiliser les questions déjà générées (sans re-génération)",
    )
    args = parser.parse_args()

    system = AblationQuerySystem()
    system.run(generate_only=args.generate_only, query_only=args.query_only)


if __name__ == "__main__":
    main()