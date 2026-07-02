"""
Calcul du BERTScore entre RAG Classique, RAG Neurosymbolique et Réponse de référence.

Installe les dépendances :
    pip install bert-score openpyxl pandas torch

Lancement :
    python bert_score_rag.py

Le fichier d'entrée et de sortie sont configurés dans la section CONFIG ci-dessous.
"""

import warnings
import os
import pandas as pd
import numpy as np
from bert_score import score as bert_score
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ===========================================================================
# ██████████████████████████  CONFIGURATION  █████████████████████████████
# ===========================================================================

# Chemin complet vers le fichier Excel source
INPUT_FILE = os.path.join(
    "C:\\", "Users", "Stage", "Desktop", "stage_RAG",
    "version_1_Ontologie", "results", "batch_queries",
    "rejet_dhydrocarbures_resultats_run49_final.xlsx"
)

# Fichier de sortie — sera créé dans le même dossier que INPUT_FILE
OUTPUT_FILE = os.path.join(
    os.path.dirname(INPUT_FILE),
    "rejet_dhydrocarbures_bertscore.xlsx"
)

# Modèle à utiliser pour le BERTScore.
# Options recommandées pour le français :
#   "camembert-base"                  → léger, rapide     (~440 Mo)  num_layers=12
#   "camembert/camembert-large"       → plus précis       (~1.3 Go)  num_layers=17
#   "microsoft/mdeberta-v3-base"      → multilingue précis (~900 Mo) num_layers=8
#   None                              → sélection automatique par langue
MODEL_TYPE = "camembert/camembert-large"

# Nombre de couches du modèle à utiliser pour le score.
# Valeurs standard : camembert-base → 12 | camembert-large → 17 | mdeberta-v3-base → 8
# Mettre None pour laisser bert_score choisir (fonctionne uniquement sur les modèles connus)
NUM_LAYERS = 17

# Langue (utilisée seulement si MODEL_TYPE est None)
LANG = "fr"

# Taille de batch — augmenter si la VRAM le permet (32, 64…), réduire si OOM
BATCH_SIZE = 16

# Utilisation du GPU — True = CUDA automatique, False = CPU forcé
USE_GPU = True

# ===========================================================================

import torch
_DEVICE = None

def _get_device():
    global _DEVICE
    if _DEVICE is None:
        if USE_GPU and torch.cuda.is_available():
            _DEVICE = "cuda"
            gpu_name = torch.cuda.get_device_name(0)
            vram     = torch.cuda.get_device_properties(0).total_memory / 1024**3
            print(f"🎮  GPU détecté : {gpu_name}  ({vram:.1f} Go VRAM)")
        else:
            _DEVICE = "cpu"
            if USE_GPU:
                print("⚠️   Aucun GPU CUDA détecté — calcul sur CPU (plus lent).")
                print("    → Pour activer le GPU, installe PyTorch CUDA :")
                print("      pip install torch --index-url https://download.pytorch.org/whl/cu121")
            else:
                print("💻  Mode CPU forcé (USE_GPU=False).")
    return _DEVICE

COL_RAG_C = "RAG Classique (dense pur)"
COL_RAG_N = "RAG Neurosymbolique (dense + sparse + graph)"
COL_REF   = "Réponse"


# ---------------------------------------------------------------------------
# Calcul
# ---------------------------------------------------------------------------

def compute_bert_scores(candidates, references):
    """
    Calcule P, R, F1 BERTScore pour une liste de (candidats, références).
    Retourne les scores moyens et par paire.
    """
    kwargs = dict(
        cands=candidates,
        refs=references,
        batch_size=BATCH_SIZE,
        device=_get_device(),
        verbose=False,
    )
    if MODEL_TYPE:
        kwargs["model_type"] = MODEL_TYPE
        if NUM_LAYERS is not None:
            kwargs["num_layers"] = NUM_LAYERS
    else:
        kwargs["lang"] = LANG

    P, R, F1 = bert_score(**kwargs)
    return {
        "P":      P.numpy().tolist(),
        "R":      R.numpy().tolist(),
        "F1":     F1.numpy().tolist(),
        "mean_P":  float(P.mean()),
        "mean_R":  float(R.mean()),
        "mean_F1": float(F1.mean()),
    }


def compute_all(xl: dict) -> pd.DataFrame:
    rows = []

    for country, df in xl.items():
        df = df.dropna(subset=[COL_REF])
        rag_c = df[COL_RAG_C].fillna("").tolist()
        rag_n = df[COL_RAG_N].fillna("").tolist()
        ref   = df[COL_REF].fillna("").tolist()
        n = len(df)

        print(f"  [{country}] {n} questions …", flush=True)

        sc_c_ref = compute_bert_scores(rag_c, ref)
        sc_n_ref = compute_bert_scores(rag_n, ref)
        sc_c_n   = compute_bert_scores(rag_c, rag_n)

        rows.append({
            "Pays":         country,
            "Nb questions": n,
            # Classique vs Référence
            "P  – Classique / Réf":  round(sc_c_ref["mean_P"],  4),
            "R  – Classique / Réf":  round(sc_c_ref["mean_R"],  4),
            "F1 – Classique / Réf":  round(sc_c_ref["mean_F1"], 4),
            # Neuro vs Référence
            "P  – Neuro / Réf":      round(sc_n_ref["mean_P"],  4),
            "R  – Neuro / Réf":      round(sc_n_ref["mean_R"],  4),
            "F1 – Neuro / Réf":      round(sc_n_ref["mean_F1"], 4),
            # Classique vs Neuro
            "P  – Classique / Neuro": round(sc_c_n["mean_P"],  4),
            "R  – Classique / Neuro": round(sc_c_n["mean_R"],  4),
            "F1 – Classique / Neuro": round(sc_c_n["mean_F1"], 4),
            # scores bruts par question (colonnes cachées, préfixe _)
            "_F1_c_ref": sc_c_ref["F1"],
            "_F1_n_ref": sc_n_ref["F1"],
            "_F1_c_n":   sc_c_n["F1"],
        })

    df_out = pd.DataFrame(rows)

    # Ligne globale
    num_cols = [c for c in df_out.columns if c.startswith(("P ", "R ", "F1"))]
    global_row = {
        "Pays": "GLOBAL (moyenne)",
        "Nb questions": df_out["Nb questions"].sum(),
        **{c: round(df_out[c].mean(), 4) for c in num_cols},
    }
    df_out = pd.concat([df_out, pd.DataFrame([global_row])], ignore_index=True)
    return df_out


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

def _thin():
    s = Side(style="thin", color="FF9DC3E6")
    return Border(top=s, bottom=s, left=s, right=s)

def _medium_top():
    m = Side(style="medium", color="FF2F75B6")
    s = Side(style="thin",   color="FF9DC3E6")
    return Border(top=m, bottom=s, left=s, right=s)


def export_excel(df: pd.DataFrame, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BERTScore"

    # Ligne 1 — titre principal
    ws.merge_cells("A1:K1")
    ws["A1"] = "BERT SCORE — RAG Classique vs RAG Neurosymbolique vs Réponse de Référence"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFFFF", name="Arial")
    ws["A1"].fill      = PatternFill("solid", start_color="FF1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Ligne 2 — info modèle
    ws.merge_cells("A2:K2")
    model_info = MODEL_TYPE if MODEL_TYPE else f"auto (lang={LANG})"
    ws["A2"] = f"Modèle : {model_info}   |   Fichier source : {os.path.basename(INPUT_FILE)}"
    ws["A2"].font      = Font(italic=True, size=9, color="FFFFFFFF", name="Arial")
    ws["A2"].fill      = PatternFill("solid", start_color="FF2E4057")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Ligne 3 — groupes
    groups = [
        ("A3:B3", ""),
        ("C3:E3", "Classique  vs  Référence"),
        ("F3:H3", "Neuro  vs  Référence"),
        ("I3:K3", "Classique  vs  Neuro"),
    ]
    grp_colors = ["FF1F3864", "FF375623", "FF1F4E79", "FF7F3F00"]
    for (rng, label), color in zip(groups, grp_colors):
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value     = label
        c.font      = Font(bold=True, size=10, color="FFFFFFFF", name="Arial")
        c.fill      = PatternFill("solid", start_color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    # Ligne 4 — en-têtes colonnes
    headers = ["Pays", "Nb\nquestions", "P", "R", "F1", "P", "R", "F1", "P", "R", "F1"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font      = Font(bold=True, size=9, color="FFFFFFFF", name="Arial")
        cell.fill      = PatternFill("solid", start_color="FF2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 36

    # Données
    data_cols = [c for c in df.columns if not c.startswith("_")]
    for r_idx, row in df.iterrows():
        er        = r_idx + 5
        is_global = row["Pays"] == "GLOBAL (moyenne)"
        bg        = "FF1B4F72" if is_global else ("FFEAF0FB" if r_idx % 2 == 0 else "FFF2F7FD")

        for c_idx, col in enumerate(data_cols, 1):
            val  = row[col]
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font      = Font(bold=is_global, size=10,
                                  color="FFFFFFFF" if is_global else "FF000000", name="Arial")
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(
                horizontal="center" if c_idx > 1 else "left", vertical="center"
            )
            if any(col.startswith(p) for p in ("P ", "R ", "F1")):
                cell.number_format = "0.0000"
            cell.border = _medium_top() if is_global else _thin()

        # Coloriser le meilleur F1 entre Classique et Neuro
        if not is_global:
            f1_c = row.get("F1 – Classique / Réf", 0) or 0
            f1_n = row.get("F1 – Neuro / Réf",     0) or 0
            best_col  = 5 if f1_c >= f1_n else 8
            worst_col = 8 if f1_c >= f1_n else 5
            ws.cell(row=er, column=best_col).fill  = PatternFill("solid", start_color="FFE2EFDA")
            ws.cell(row=er, column=worst_col).fill = PatternFill("solid", start_color="FFFCE4D6")

        ws.row_dimensions[er].height = 18

    # Largeurs colonnes
    for i, w in enumerate([26, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Légende
    lr = len(df) + 7
    ws.merge_cells(f"A{lr}:K{lr}")
    ws[f"A{lr}"] = (
        "P = Précision  |  R = Rappel  |  F1 = F-mesure  "
        "|  🟢 Meilleur F1 entre Classique et Neuro  "
        "|  🔴 Score inférieur  "
        "|  BERTScore ∈ [0, 1] — plus élevé = plus similaire sémantiquement"
    )
    ws[f"A{lr}"].font      = Font(italic=True, size=9, color="FF595959", name="Arial")
    ws[f"A{lr}"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[lr].height = 30

    ws.freeze_panes = "A5"
    wb.save(path)
    print(f"\n✅  Résultats sauvegardés dans : {path}")


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("=" * 60)
    print("  BERTScore — Évaluation des systèmes RAG")
    print("=" * 60)
    print(f"\n📂  Fichier source  : {INPUT_FILE}")
    print(f"💾  Fichier sortie  : {OUTPUT_FILE}")
    print(f"🤖  Modèle          : {MODEL_TYPE or f'auto (lang={LANG})'}")
    print(f"📦  Batch size      : {BATCH_SIZE}")
    _get_device()
    print()

    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(
            f"Fichier introuvable :\n  {INPUT_FILE}\n"
            "Vérifiez le chemin dans la section CONFIG du script."
        )

    print("📖  Chargement du fichier Excel …")
    xl = pd.read_excel(INPUT_FILE, sheet_name=None)
    print(f"   {len(xl)} pays trouvés : {', '.join(xl.keys())}\n")

    print("🔢  Calcul des BERTScores (peut prendre quelques minutes) …\n")
    df_results = compute_all(xl)

    print("\n📊  Résumé F1 :")
    print("-" * 60)
    summary = df_results[["Pays", "F1 – Classique / Réf", "F1 – Neuro / Réf", "F1 – Classique / Neuro"]]
    print(summary.to_string(index=False))

    print("\n💾  Export Excel …")
    export_excel(df_results, OUTPUT_FILE)