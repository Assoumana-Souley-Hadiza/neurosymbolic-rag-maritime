"""
Script d'étude d'ablation du système RAG
Compare 4 configurations de pipeline pour mesurer l'impact de chaque composant.

Architecture confirmée (identique à batch_query.py) :
  - Le Graph Retriever N'EST PAS utilisé côté retrieval.
  - Le Neo4j sert uniquement à :
      (a) l'expansion de requête  → expanded_terms / synonym_sets  (dense + sparse)
      (b) l'enrichissement LLM    → ontology_agent.enrich()

Les 4 pipelines isolent donc :

  ┌─────────────────┬────────┬────────┬────────────────┬────────────────────┐
  │ Pipeline        │ Dense  │ Sparse │ Expansion Neo4j│ Enrichiss. Ontolog.│
  ├─────────────────┼────────┼────────┼────────────────┼────────────────────┤
  │ Classique       │ brut   │  ✗     │      ✗         │        ✗           │
  │ Sans_Sparse     │ enrichi│  ✗     │      ✓         │        ✓           │
  │ Sans_Ontologie  │ enrichi│  ✓     │      ✓         │        ✗           │
  │ Complet         │ enrichi│  ✓     │      ✓         │        ✓           │
  └─────────────────┴────────┴────────┴────────────────┴────────────────────┘

Comparaisons utiles :
  Classique     → Sans_Sparse    : apport de l'expansion Neo4j + ontologie seuls
  Sans_Sparse   → Complet        : apport du sparse retriever
  Sans_Ontologie→ Complet        : apport de l'enrichissement ontologique
  Classique     → Complet        : gain global du pipeline neuro
"""

import sys
import csv
import json
import logging
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import argparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from rag.core.fusion import HybridFusion
from rag.core.query_analyzer import QueryAnalyzer
from rag.llm_generator import LLMGenerator
from rag.integration.neo4j_bridge import Neo4jBridge
from rag.core.retrievers import DenseRetriever, SparseRetriever
from rag.neo4j_ontology_agent import Neo4jOntologyAgent

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('batch_ablation_ontologie.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# PAYS
# ─────────────────────────────────────────────────────────────────────────────

PAYS = {
    "Bénin":         {"code": "ben", "prep": "au Bénin",          "prep_q": "au Bénin"},
    "Cameroun":      {"code": "cmr", "prep": "au Cameroun",       "prep_q": "au Cameroun"},
    "Comores":       {"code": "com", "prep": "aux Comores",       "prep_q": "aux Comores"},
    "Congo":         {"code": "cng", "prep": "au Congo",          "prep_q": "au Congo"},
    "Côte d'Ivoire": {"code": "ivc", "prep": "en Côte d'Ivoire", "prep_q": "en Côte d'Ivoire"},
    "Djibouti":      {"code": "dji", "prep": "à Djibouti",        "prep_q": "à Djibouti"},
    "Gabon":         {"code": "gab", "prep": "au Gabon",          "prep_q": "au Gabon"},
    "Guinée":        {"code": "gui", "prep": "en Guinée",         "prep_q": "en Guinée"},
    "Madagascar":    {"code": "mad", "prep": "à Madagascar",      "prep_q": "à Madagascar"},
    "Maroc":         {"code": "mor", "prep": "au Maroc",          "prep_q": "au Maroc"},
    "Mauritanie":    {"code": "mau", "prep": "en Mauritanie",     "prep_q": "en Mauritanie"},
    "Sénégal":       {"code": "sen", "prep": "au Sénégal",        "prep_q": "au Sénégal"},
    "Togo":          {"code": "tog", "prep": "au Togo",           "prep_q": "au Togo"},
    "Tunisie":       {"code": "tun", "prep": "en Tunisie",        "prep_q": "en Tunisie"},
}


# ─────────────────────────────────────────────────────────────────────────────
# QUESTIONS
# ─────────────────────────────────────────────────────────────────────────────

INTERDICTIONS = {
    "Chalutage de fond": [
        "Est-il interdit de pratiquer la pêche au chalut {prep} ?",
        "A-t-on le droit de faire du dragage des fonds marins {prep} ?",
        "Le chalutage benthique est-il autorisé {prep} ?"
    ],
    "Chasse à la baleine": [
        "Peut-on chasser les cétacés {prep} ?",
        "La capture des mammifères marins est-elle interdite {prep} ?",
        "Quelles sont les règles sur l'abattage des mammifères aquatiques {prep} ?"
    ],
    "Construction côtière": [
        "A-t-on le droit de procéder à une édification sur le rivage {prep} ?",
        "Est-il permis de faire des aménagements sur le littoral {prep} ?",
        "Peut-on bâtir sur la zone côtière {prep} ?"
    ],
    "Oiseaux Marins": [
        "A-t-on le droit de chasser l'avifaune marine {prep} ?",
        "Est-il interdit de capturer des oiseaux de mer {prep} ?",
        "Que dit la loi sur la détention d'espèces aviaires marines {prep} ?"
    ],
    "Rejet d'hydrocarbures": [
        "Est-il permis de faire un déversement de pétrole {prep} ?",
        "Est-il interdit aux navires de faire un rejet de fioul {prep} ?",
        "Peut-on vider les huiles usagées en mer {prep} ?"
    ]
}



# ─────────────────────────────────────────────────────────────────────────────
# DÉFINITION DES PIPELINES D'ABLATION
# ─────────────────────────────────────────────────────────────────────────────
#
#  use_sparse    : active le SparseRetriever dans la fusion
#  use_expansion : active l'expansion de requête Neo4j (expanded_terms + synonym_sets)
#                  Elle s'applique à la fois sur le Dense ET le Sparse quand actif.
#  use_ontology  : active l'enrichissement de contexte via Neo4jOntologyAgent.enrich()
#
#  NOTE : il n'y a PAS de "use_graph" — le graph retriever n'est jamais utilisé
#         côté retrieval (cohérent avec batch_query.py et l'archi Streamlit).

ABLATION_PIPELINES: List[Dict] = [
    {
        "name":          "Classique",
        "label":         "RAG Classique\n(dense pur)",
        "use_sparse":    False,
        "use_expansion": False,  # Dense brut, sans expanded_terms ni synonym_sets
        "use_ontology":  False,  # enriched_context = None
        "header_color":  "2F4F8F",  # bleu marine
    },
    {
        "name":          "Complet",
        "label":         "RAG Complet\n(dense enrichi + sparse + ontologie)",
        "use_sparse":    True,
        "use_expansion": True,
        "use_ontology":  True,
        "header_color":  "4A7C59",  # vert sombre
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# UTILITAIRES EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def _slugify(text: str) -> str:
    normalized = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('utf-8')
    return normalized.lower().replace(" ", "_").replace("'", "")


def _pays_sheet_name(pays_nom: str) -> str:
    forbidden = r"\/*?:[]"
    name = pays_nom
    for ch in forbidden:
        name = name.replace(ch, "_")
    return name[:31]


def _clean_for_excel(val) -> str:
    if not isinstance(val, str):
        return val
    val = re.sub(r'[\x01-\x08]', '• ', val)
    return re.sub(r'[\x00\x0b\x0c\x0e-\x1f]', '', val)


_HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN   = Alignment(vertical="top", wrap_text=True)
_BORDER_THIN  = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
_ALT_FILL = PatternFill("solid", start_color="F2F2F2")


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color)


def _style_sheet(ws, col_widths: List[int], header_fills: List[PatternFill]):
    for col_idx, (cell, fill) in enumerate(zip(ws[1], header_fills), 1):
        cell.font      = _HEADER_FONT
        cell.fill      = fill
        cell.alignment = _HEADER_ALIGN
        cell.border    = _BORDER_THIN

    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for cell in row:
            cell.alignment = _CELL_ALIGN
            cell.border    = _BORDER_THIN
            if fill:
                cell.fill = fill

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 80


# ─────────────────────────────────────────────────────────────────────────────
# CLASSE PRINCIPALE
# ─────────────────────────────────────────────────────────────────────────────

class AblationQuerySystem:
    """
    Orchestre l'étude d'ablation en exécutant chaque pipeline RAG séparément.

    Principe :
      - L'expansion Neo4j (synonymes, expanded_terms) est calculée UNE SEULE FOIS
        par question, puis transmise ou non à chaque pipeline selon sa config.
      - Le Graph Retriever n'est JAMAIS utilisé dans la fusion :
        seul l'OntologyAgent peut l'utiliser en interne pour l'enrichissement.
      - Le pipeline Classique reproduit exactement le comportement de
        batch_query.py / RAG Classique : dense pur, aucun enrichissement.
      - Le pipeline Complet reproduit exactement batch_query.py / RAG Neuro.
    """

    # Mots de bruit et keywords anglais désormais gérés par l'OntologyAgent.

    def __init__(self):
        self.output_dir = Path("results/ablation_study")
        self.output_dir.mkdir(parents=True, exist_ok=True)

        existing_runs = []
        for f in self.output_dir.glob("*_run*.xlsx"):
            try:
                existing_runs.append(int(f.stem.split('_run')[-1]))
            except ValueError:
                pass
        self.run_id = max(existing_runs) + 1 if existing_runs else 1

        self.run_timestamp  = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.questions_file = self.output_dir / "questions_generees.csv"
        self.results_csv    = self.output_dir / f"ablation_onto_res_run{self.run_id}_{self.run_timestamp}.csv"
        self.summary_file   = self.output_dir / f"ablation_onto_res_run{self.run_id}_{self.run_timestamp}.json"

        # Composants RAG (initialisés dans init_rag_system)
        self.fusion           = None
        self.llm_gen          = None
        self.query_analyzer   = None
        self.neo4j_bridge     = None
        self.ontology_agent   = None
        self.dense_retriever  = None
        self.sparse_retriever = None

        logger.info(f"AblationQuerySystem — run #{self.run_id}")

    # ─────────────────────────────────────────────────────────────────────────
    # Initialisation
    # ─────────────────────────────────────────────────────────────────────────

    def init_rag_system(self) -> bool:
        try:
            logger.info("Initialisation des composants RAG...")

            self.dense_retriever  = DenseRetriever()
            self.sparse_retriever = SparseRetriever()

            self.neo4j_bridge = Neo4jBridge.from_config()
            if not self.neo4j_bridge.is_ready():
                logger.warning("⚠ Neo4j Bridge non connecté — expansion et ontologie seront ignorés.")
            else:
                logger.info("✓ Neo4j Bridge connecté")

            # OntologyAgent uniquement (pas de GraphRetriever en retrieval)
            self.ontology_agent = Neo4jOntologyAgent(self.neo4j_bridge)

            self.fusion         = HybridFusion()
            self.query_analyzer = QueryAnalyzer()
            self.llm_gen        = LLMGenerator()

            if not self.llm_gen.is_ollama_available():
                logger.warning("⚠ LLM non disponible")
                return False

            logger.info("✓ Tous les composants RAG prêts")
            return True

        except Exception as e:
            logger.error(f"❌ Erreur d'initialisation : {e}", exc_info=True)
            return False

    # ─────────────────────────────────────────────────────────────────────────
    # Génération des questions
    # ─────────────────────────────────────────────────────────────────────────

    def generer_questions(self) -> List[Dict]:
        questions = []
        for pays_nom, pays_info in PAYS.items():
            prep, prep_q = pays_info["prep"], pays_info["prep_q"]
            for interdiction, templates in INTERDICTIONS.items():
                for num_q, template in enumerate(templates, 1):
                    questions.append({
                        "Pays":            pays_nom,
                        "Code_Pays":       pays_info["code"],
                        "Interdiction":    interdiction,
                        "Numero_Question": num_q,
                        "Question":        template.replace("{prep}", prep).replace("{prep_q}", prep_q),
                    })

        with open(self.questions_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=questions[0].keys())
            writer.writeheader()
            writer.writerows(questions)

        print(f"✓ {len(questions)} questions générées → {self.questions_file}")
        return questions

    # ─────────────────────────────────────────────────────────────────────────
    # Expansion Neo4j — calculée UNE FOIS par question, partagée entre pipelines
    # ─────────────────────────────────────────────────────────────────────────

    def _build_expansion(self, question: str) -> Tuple[Dict, List[str], set]:
        """
        Interroge Neo4j pour obtenir :
          - synonym_sets   : dict {interdiction → set(synonymes)}
          - expanded_terms : liste plate des synonymes (pour dense/sparse)
          - graph_entities : labels d'entités du graphe (pour ontology_boost)

        Retourne des structures vides si Neo4j n'est pas disponible.
        Logique identique à batch_query.py.
        """
        synonym_sets:   Dict = {}
        expanded_terms: List = []
        graph_entities: set  = set()

    def _build_expansion(self, question: str) -> Tuple[Dict, List[str], set]:
        synonym_sets:   Dict = {}
        expanded_terms: List = []
        graph_entities: set  = set()

        if not (self.ontology_agent and self.ontology_agent.is_ready()):
            return synonym_sets, expanded_terms, graph_entities

        query_context = self.ontology_agent.prepare_query(question)
        synonym_sets = query_context.get("synonyms_map", {})
        graph_entities = query_context.get("graph_entities", set())
        expanded_terms = query_context.get("expanded_terms", [])

        return synonym_sets, expanded_terms, graph_entities

    # ─────────────────────────────────────────────────────────────────────────
    # Exécution d'un seul pipeline d'ablation
    # ─────────────────────────────────────────────────────────────────────────

    def _run_pipeline(
        self,
        pipeline:        Dict,
        question:        str,
        intent:          str,
        weights:         Dict,
        country_filter,
        category_filter,
        synonym_sets:    Dict,
        expanded_terms:  List[str],
        graph_entities:  set,
    ) -> str:
        """
        Exécute un pipeline d'ablation et retourne la réponse LLM nettoyée.

        Règle fondamentale :
          - Le Graph Retriever n'est JAMAIS appelé ici.
          - use_expansion contrôle les expanded_terms passés aux retrievers
            ET les synonym_sets passés à la fusion.
          - use_ontology contrôle uniquement ontology_agent.enrich().
        """
        _K = 12

        # ── Retrievers ────────────────────────────────────────────────────────

        retriever_results: Dict = {}

        # Dense — avec ou sans expanded_terms selon le pipeline
        if self.dense_retriever and self.dense_retriever.is_ready():
            dense_kwargs = {"top_k": _K}
            if pipeline["use_expansion"] and expanded_terms:
                dense_kwargs["expanded_terms"] = expanded_terms
            retriever_results["dense"] = self.dense_retriever.retrieve(
                question, **dense_kwargs
            )

        # Sparse — activé uniquement si use_sparse=True
        if pipeline["use_sparse"] and self.sparse_retriever and self.sparse_retriever.is_ready():
            sparse_kwargs = {"top_k": _K}
            if pipeline["use_expansion"] and expanded_terms:
                sparse_kwargs["expanded_terms"] = expanded_terms
            retriever_results["sparse"] = self.sparse_retriever.retrieve(
                question, **sparse_kwargs
            )

        # ── Fusion ────────────────────────────────────────────────────────────

        fusion_kwargs: Dict = {
            "retriever_results": retriever_results,
            "weights":           weights,
            "synonym_sets":      synonym_sets if pipeline["use_expansion"] else {},
            "top_k":             6,
            "country_filter":    country_filter,
            "query":             question,
            "disable_reranker":  pipeline["name"] == "Classique",
        }

        # graph_entities et ontology_boost n'ont de sens qu'avec expansion activée
        if pipeline["use_expansion"] and graph_entities:
            fusion_kwargs["graph_entities"] = graph_entities
            fusion_kwargs["ontology_boost"] = 0.5

        if pipeline["use_expansion"] and category_filter:
            fusion_kwargs["category_filter"] = category_filter

        if pipeline["name"] == "Classique":
            # RAG CLASSIQUE STRICT: On prend juste le top_k=6 du dense, sans passer par la fusion
            fusion_results = retriever_results["dense"][:6]
        else:
            fusion_results = self.fusion.fuse(**fusion_kwargs)
        # ── Enrichissement ontologique (optionnel) ────────────────────────────

        enriched_context = None
        if pipeline["use_ontology"] and self.ontology_agent and self.ontology_agent.is_ready():
            try:
                enriched_context = self.ontology_agent.enrich(question, fusion_results)
            except Exception as e:
                logger.warning(f"[{pipeline['name']}] Ontology agent ignoré : {e}")

        # ── Génération LLM ────────────────────────────────────────────────────

        response_gen = self.llm_gen.generate(
            query=question,
            fusion_results=fusion_results,
            intent=intent,
            enriched_context=enriched_context,
            stream=True,
        )

        if response_gen.get("error"):
            return "ERREUR"

        full_ans = "".join(list(response_gen["stream_generator"]))
        return re.sub(r'<analyse>.*?(?:</analyse>|$)', '', full_ans, flags=re.DOTALL).strip()

    # ─────────────────────────────────────────────────────────────────────────
    # Traitement d'une question (tous pipelines)
    # ─────────────────────────────────────────────────────────────────────────

    def interroger_question(self, question_data: Dict) -> Dict:
        """
        Pour une question donnée :
          1. Analyse commune (intent, weights, country_filter, category_filter)
          2. Expansion Neo4j UNE SEULE FOIS (shared entre pipelines)
          3. Exécution séquentielle des 4 pipelines d'ablation
        """
        question = question_data["Question"]

        base = {
            "Pays":            question_data["Pays"],
            "Code_Pays":       question_data["Code_Pays"],
            "Interdiction":    question_data["Interdiction"],
            "Numero_Question": question_data["Numero_Question"],
            "Question":        question,
            "Modele":          "N/A",
            "Timestamp":       datetime.now().isoformat(),
            "Erreur":          "",
        }
        # Pré-remplissage des colonnes réponses
        for p in ABLATION_PIPELINES:
            base[f"Reponse_{p['name']}"] = "ERREUR"

        try:
            if not self.query_analyzer or not self.fusion or not self.llm_gen:
                raise RuntimeError("Système RAG non initialisé")

            base["Modele"] = self.llm_gen.model_name

            # ── 1. Analyse commune ─────────────────────────────────────────
            intent, weights, country_filter, category_filter = \
                self.query_analyzer.analyze(question)

            # ── 2. Expansion Neo4j (calculée une seule fois) ───────────────
            # On ne calcule que si au moins un pipeline actif l'utilise.
            # Le pipeline Classique (use_expansion=False) n'a pas besoin de Neo4j.
            needs_expansion = any(p["use_expansion"] for p in ABLATION_PIPELINES)
            if needs_expansion:
                synonym_sets, expanded_terms, graph_entities = \
                    self._build_expansion(question)
            else:
                synonym_sets, expanded_terms, graph_entities = {}, [], set()

            # ── 3. Exécution de chaque pipeline ───────────────────────────
            errors = []
            for pipeline in ABLATION_PIPELINES:
                try:
                    answer = self._run_pipeline(
                        pipeline        = pipeline,
                        question        = question,
                        intent          = intent,
                        weights         = weights,
                        country_filter  = country_filter,
                        category_filter = category_filter,
                        synonym_sets    = synonym_sets,
                        expanded_terms  = expanded_terms,
                        graph_entities  = graph_entities,
                    )
                    base[f"Reponse_{pipeline['name']}"] = answer
                    logger.debug(f"[{pipeline['name']}] ✓")
                except Exception as e:
                    logger.error(
                        f"[{pipeline['name']}] Erreur pipeline : {e}", exc_info=True
                    )
                    errors.append(f"{pipeline['name']}: {e}")

            if errors:
                base["Erreur"] = " | ".join(errors)

        except Exception as e:
            logger.error(f"Erreur traitement question : {e}", exc_info=True)
            base["Erreur"] = str(e)

        return base

    # ─────────────────────────────────────────────────────────────────────────
    # Sauvegarde CSV
    # ─────────────────────────────────────────────────────────────────────────

    def sauvegarder_resultats(self, resultats: List[Dict]):
        if not resultats:
            return
        fieldnames = [
            "Pays", "Code_Pays", "Interdiction", "Numero_Question", "Question",
            *[f"Reponse_{p['name']}" for p in ABLATION_PIPELINES],
            "Modele", "Erreur", "Timestamp",
        ]
        with open(self.results_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(resultats)

    # ─────────────────────────────────────────────────────────────────────────
    # Excel — écriture incrémentale
    # ─────────────────────────────────────────────────────────────────────────

    def mettre_a_jour_excel(self, resultats: List[Dict]):
        """Recrée les fichiers Excel par interdiction à chaque appel (incrémental)."""
        par_interdiction: Dict[str, List[Dict]] = {}
        for r in resultats:
            par_interdiction.setdefault(r["Interdiction"], []).append(r)

        for interdiction, rows in par_interdiction.items():
            slug       = _slugify(interdiction)
            excel_path = self.output_dir / f"ablation_onto_{slug}_run{self.run_id}.xlsx"

            rows_by_pays: Dict[str, List[Dict]] = {}
            for r in rows:
                rows_by_pays.setdefault(r["Pays"], []).append(r)

            self._ecrire_excel(excel_path, rows_by_pays)

    def _ecrire_excel(self, output_path: Path, rows_by_pays: Dict[str, List[Dict]]):
        """Crée un fichier Excel comparatif avec une feuille par pays + synthèse."""
        wb = Workbook()
        if wb.active is not None:
            wb.remove(wb.active)

        # En-têtes et couleurs basés sur ABLATION_PIPELINES
        headers = ["Question"] + [p["label"] for p in ABLATION_PIPELINES]
        fills   = [_make_fill("2F4F4F")] + [_make_fill(p["header_color"]) for p in ABLATION_PIPELINES]

        n_pipelines        = len(ABLATION_PIPELINES)
        response_col_width = max(10, 220 // n_pipelines)
        col_widths         = [55] + [response_col_width] * n_pipelines

        for pays_nom in PAYS.keys():
            rows = rows_by_pays.get(pays_nom, [])
            if not rows:
                continue

            ws = wb.create_sheet(title=_pays_sheet_name(pays_nom))
            ws.append(headers)
            ws.row_dimensions[1].height = 40

            for r in rows:
                row_data = [_clean_for_excel(r["Question"])]
                for p in ABLATION_PIPELINES:
                    row_data.append(_clean_for_excel(r.get(f"Reponse_{p['name']}", "N/A")))
                ws.append(row_data)

            _style_sheet(ws, col_widths=col_widths, header_fills=fills)

        # Feuille de synthèse OUI/NON
        self._add_summary_sheet(wb, rows_by_pays)

        wb.save(output_path)
        logger.info(f"✓ Excel sauvegardé : {output_path}")

    def _add_summary_sheet(self, wb: Workbook, rows_by_pays: Dict[str, List[Dict]]):
        """Feuille 'Synthèse' : taux OUI/NON par pays et par pipeline."""
        ws = wb.create_sheet(title="Synthèse", index=0)

        headers = ["Pays", "Total"]
        for p in ABLATION_PIPELINES:
            headers += [f"OUI — {p['name']}", f"NON — {p['name']}"]
        ws.append(headers)

        all_rows: List[Dict] = [r for rows in rows_by_pays.values() for r in rows]

        for pays_nom in PAYS.keys():
            pays_rows = [r for r in all_rows if r["Pays"] == pays_nom]
            if not pays_rows:
                continue
            row_data: List = [pays_nom, len(pays_rows)]
            for p in ABLATION_PIPELINES:
                col = f"Reponse_{p['name']}"
                oui = sum(1 for r in pays_rows if "OUI" in str(r.get(col, "")).upper())
                non = sum(1 for r in pays_rows if "NON" in str(r.get(col, "")).upper())
                row_data += [oui, non]
            ws.append(row_data)

        # Ligne de totaux
        total_row: List = ["TOTAL", len(all_rows)]
        for p in ABLATION_PIPELINES:
            col = f"Reponse_{p['name']}"
            total_row += [
                sum(1 for r in all_rows if "OUI" in str(r.get(col, "")).upper()),
                sum(1 for r in all_rows if "NON" in str(r.get(col, "")).upper()),
            ]
        ws.append(total_row)

        summary_fills = [_make_fill("2F4F4F"), _make_fill("2F4F4F")]
        for p in ABLATION_PIPELINES:
            summary_fills += [_make_fill(p["header_color"]), _make_fill(p["header_color"])]

        col_widths_s = [20, 8] + [16] * (2 * len(ABLATION_PIPELINES))
        _style_sheet(ws, col_widths=col_widths_s, header_fills=summary_fills)

    # ─────────────────────────────────────────────────────────────────────────
    # Résumé JSON
    # ─────────────────────────────────────────────────────────────────────────

    def generer_resume(self, resultats: List[Dict]) -> Dict:
        def _stats(rows: List[Dict]) -> Dict:
            out: Dict = {
                "total":   len(rows),
                "erreurs": sum(
                    1 for r in rows
                    if any(r.get(f"Reponse_{p['name']}") == "ERREUR" for p in ABLATION_PIPELINES)
                ),
            }
            for p in ABLATION_PIPELINES:
                col = f"Reponse_{p['name']}"
                out[f"oui_{p['name']}"] = sum(
                    1 for r in rows if "OUI" in str(r.get(col, "")).upper()
                )
                out[f"non_{p['name']}"] = sum(
                    1 for r in rows if "NON" in str(r.get(col, "")).upper()
                )
            return out

        resume = {
            "date_execution": datetime.now().isoformat(),
            "run_id":         self.run_id,
            "pipelines":      [
                {"name": p["name"], "config": {
                    "use_sparse":    p["use_sparse"],
                    "use_expansion": p["use_expansion"],
                    "use_ontology":  p["use_ontology"],
                }}
                for p in ABLATION_PIPELINES
            ],
            **_stats(resultats),
            "par_pays":         {
                pays: _stats([r for r in resultats if r["Pays"] == pays])
                for pays in PAYS
            },
            "par_interdiction": {
                inter: _stats([r for r in resultats if r["Interdiction"] == inter])
                for inter in INTERDICTIONS
            },
        }

        with open(self.summary_file, 'w', encoding='utf-8') as f:
            json.dump(resume, f, ensure_ascii=False, indent=2)

        logger.info(f"✓ Résumé JSON → {self.summary_file}")
        return resume

    def afficher_resume(self, resume: Dict):
        print("\n" + "=" * 72)
        print("📊  RÉSUMÉ DE L'ÉTUDE D'ABLATION")
        print("=" * 72)
        print(f"Total questions : {resume['total']}   |   Erreurs : {resume['erreurs']}")
        print()

        col_w = 22
        header = f"{'Pipeline':<{col_w}} {'OUI':>6} {'NON':>6}   Description"
        print(header)
        print("-" * 72)

        descriptions = {
            "Classique":      "dense pur, sans expansion ni ontologie",
            "Sans_Sparse":    "dense enrichi + ontologie, sans sparse",
            "Sans_Ontologie": "dense enrichi + sparse, sans ontologie",
            "Complet":        "dense enrichi + sparse + ontologie",
        }

        for p in ABLATION_PIPELINES:
            n   = p["name"]
            oui = resume.get(f"oui_{n}", 0)
            non = resume.get(f"non_{n}", 0)
            desc = descriptions.get(n, "")
            print(f"{n:<{col_w}} {oui:>6} {non:>6}   {desc}")

        print("=" * 72 + "\n")

    # ─────────────────────────────────────────────────────────────────────────
    # Pipeline principal
    # ─────────────────────────────────────────────────────────────────────────

    def run(self, generate_only: bool = False, query_only: bool = False):
        try:
            # ── Chargement / génération des questions ──────────────────────
            if not query_only:
                questions = self.generer_questions()
            else:
                questions = []
                with open(self.questions_file, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        row["Numero_Question"] = int(row["Numero_Question"])
                        questions.append(row)
                print(f"✓ {len(questions)} questions chargées depuis {self.questions_file}")

            if generate_only:
                print("\n✅ Génération terminée. Utilisez --query-only pour lancer les pipelines.")
                return

            # ── Initialisation du système RAG ──────────────────────────────
            if not self.init_rag_system():
                logger.error("❌ Impossible d'initialiser le système RAG")
                return

            resultats: List[Dict] = []
            total   = len(questions)
            n_pipes = len(ABLATION_PIPELINES)

            print(f"\n🔬 Étude d'ablation — {total} questions × {n_pipes} pipelines\n")
            for p in ABLATION_PIPELINES:
                flag_sparse  = "✓ sparse"    if p["use_sparse"]    else "✗ sparse"
                flag_expand  = "✓ expansion" if p["use_expansion"] else "✗ expansion"
                flag_onto    = "✓ ontologie" if p["use_ontology"]  else "✗ ontologie"
                print(f"  [{p['name']:<16}]  {flag_sparse}  {flag_expand}  {flag_onto}")
            print()

            # ── Boucle principale ──────────────────────────────────────────
            for idx, q_data in enumerate(questions, 1):
                label = (
                    f"[{idx}/{total}] {q_data['Pays']} — "
                    f"{q_data['Interdiction']} (Q{q_data['Numero_Question']})"
                )
                print(label, end=" ", flush=True)

                result = self.interroger_question(q_data)
                resultats.append(result)
                print("✓")

                # Sauvegarde incrémentale après chaque question
                self.sauvegarder_resultats(resultats)
                self.mettre_a_jour_excel(resultats)

                if idx % 10 == 0:
                    logger.info(f"Progression : {idx}/{total} questions traitées")

            # ── Résumé final ───────────────────────────────────────────────
            resume = self.generer_resume(resultats)
            self.afficher_resume(resume)
            print(f"✅ Résultats dans : {self.output_dir}")

        except Exception as e:
            logger.error(f"❌ Erreur fatale : {e}", exc_info=True)
            raise


# ─────────────────────────────────────────────────────────────────────────────
# POINT D'ENTRÉE
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Étude d'ablation du système RAG (4 pipelines comparés)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="\n".join(
            f"  {p['name']:<18} {p['label'].replace(chr(10), ' ')}"
            for p in ABLATION_PIPELINES
        ),
    )
    parser.add_argument(
        "--generate-only", action="store_true",
        help="Générer les questions uniquement (pas d'interrogation)",
    )
    parser.add_argument(
        "--query-only", action="store_true",
        help="Utiliser les questions déjà générées (sans re-génération)",
    )
    args = parser.parse_args()

    system = AblationQuerySystem()
    system.run(generate_only=args.generate_only, query_only=args.query_only)


if __name__ == "__main__":
    main()