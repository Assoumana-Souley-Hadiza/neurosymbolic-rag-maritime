"""
Script d'interrogation automatique du système RAG
Génère les questions par pays/interdiction et stocke les réponses sous formats CSV, JSON et Excel.
"""

import sys
import csv
import json
import logging
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import argparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ajouter le dossier racine au PYTHONPATH
sys.path.insert(0, str(Path(__file__).resolve().parent))

from rag.core.fusion import HybridFusion
from rag.core.query_analyzer import QueryAnalyzer
from rag.llm_generator import LLMGenerator
from rag.integration.neo4j_bridge import Neo4jBridge
from rag.core.neo4j_graph_retriever import Neo4jGraphRetriever
from rag.core.retrievers import DenseRetriever, SparseRetriever
from rag.neo4j_ontology_agent import Neo4jOntologyAgent

# Configuration du logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('batch_query.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# PAYS : code interne + forme prépositionnelle naturelle en français
# ─────────────────────────────────────────────────────────────────────────────
# "prep"  → utilisé dans "l'interdiction de la chasse des baleines {prep}"
# "prep_q"→ utilisé dans "Est-ce qu'{prep_q}, des procédures…"
#           (même valeur sauf si la phrase "qu'" + prep sonne mieux sans virgule)

PAYS = {
    "Bénin":        {"code": "ben", "prep": "au Bénin",          "prep_q": "au Bénin"},
    "Cameroun":     {"code": "cmr", "prep": "au Cameroun",       "prep_q": "au Cameroun"},
    "Comores":      {"code": "com", "prep": "aux Comores",       "prep_q": "aux Comores"},
    "Congo":        {"code": "cng", "prep": "au Congo",          "prep_q": "au Congo"},
    "Côte d'Ivoire":{"code": "ivc", "prep": "en Côte d'Ivoire", "prep_q": "en Côte d'Ivoire"},
    "Djibouti":     {"code": "dji", "prep": "à Djibouti",        "prep_q": "à Djibouti"},
    "Gabon":        {"code": "gab", "prep": "au Gabon",          "prep_q": "au Gabon"},
    "Guinée":       {"code": "gui", "prep": "en Guinée",         "prep_q": "en Guinée"},
    "Madagascar":   {"code": "mad", "prep": "à Madagascar",      "prep_q": "à Madagascar"},
    "Maroc":        {"code": "mor", "prep": "au Maroc",          "prep_q": "au Maroc"},
    "Mauritanie":   {"code": "mau", "prep": "en Mauritanie",     "prep_q": "en Mauritanie"},
    "Sénégal":      {"code": "sen", "prep": "au Sénégal",        "prep_q": "au Sénégal"},
    "Togo":         {"code": "tog", "prep": "au Togo",           "prep_q": "au Togo"},
    "Tunisie":      {"code": "tun", "prep": "en Tunisie",        "prep_q": "en Tunisie"},
}

# ─────────────────────────────────────────────────────────────────────────────
# QUESTIONS : {prep} = forme naturelle, {prep_q} = forme après "qu'"
# ─────────────────────────────────────────────────────────────────────────────

INTERDICTIONS = {
    "Chasse à la baleine": [
        "Est-ce qu'il existe un article portant sur l'interdiction de la chasse des baleines {prep} ?",
        "Est-ce qu'il existe un article portant sur l'interdiction de la chasse des baleines {prep} qui précise que cette mesure n'est applicable qu'à certaines zones, aires ou régions ?",
        "Est-ce qu'il existe un article portant sur l'interdiction de la chasse des baleines {prep} qui précise que cette mesure n'est pas applicable en permanence (temporalité) ?",
        "Est-ce qu'un texte juridique précise les types d'activités concernées par l'interdiction de la chasse des baleines {prep} ?",
        "Est-ce qu'un texte juridique mentionne des exceptions à l'interdiction de la chasse des baleines {prep} ? Vérifie si ce texte inclut des exceptions relatives à d'autres domaines que de la santé, de l'ordre public et de la recherche.",
        "Est-ce qu'un texte juridique mentionne des exceptions à l'interdiction de la chasse des baleines {prep} ? Ne pas inclure les délais de mise en place de la loi. Les dérogations ou autorisations spécifiques constituent des exceptions.",
        "Est-ce qu'il existe un article qui précise que l'infraction à l'interdiction de la chasse des baleines {prep} entraîne une sanction financière ? ",
        "Est-ce que l'infraction à l'interdiction de la chasse des baleines {prep} entraîne une peine de prison ?",
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction de la chasse des baleines ?",
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction de la chasse des baleines ? Si ces procédures existent, vérifiez si elles précisent des périodes (temporalité) spécifiques pour le contrôle du respect de l'interdiction de la chasse des baleines.",
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction de la chasse des baleines ? Si ces procédures existent, vérifiez si elles précisent des lieux ou zones spécifiques pour le contrôle du respect de l'interdiction de la chasse des baleines.",
    ],
    "Rejet d'hydrocarbures": [
        "Est-ce qu'il existe un article portant sur l'interdiction du rejet d'hydrocarbures {prep} ?",
        "Est-ce qu'il existe un article portant sur l'interdiction du rejet d'hydrocarbures {prep} qui précise que cette mesure n'est applicable qu'à certaines zones, aires ou régions ?",
        "Est-ce qu'il existe un article portant sur l'interdiction du rejet d'hydrocarbures {prep} qui précise que cette mesure n'est pas applicable en permanence (temporalité) ?",
        "Est-ce que le texte juridique précise les types d'activités concernées par l'interdiction du rejet d'hydrocarbures {prep} ?",
        "Est-ce que le texte juridique mentionne des exceptions à l'interdiction du rejet d'hydrocarbures {prep} ? Vérifie si ce texte inclut des exceptions relatives à d'autres domaines que de la santé, de l'ordre public et de la recherche.",
        "Est-ce que le texte juridique mentionne des exceptions à l'interdiction du rejet d'hydrocarbures {prep} ?",
        "Est-ce que le rejet d'hydrocarbures {prep} entraîne une sanction financière (amende) ?",
        "Est-ce que le rejet d'hydrocarbures {prep} entraîne une peine de prison ?",
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction du rejet d'hydrocarbures ?",
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction du rejet d'hydrocarbures ? Si ces procédures existent, vérifiez si elles précisent des périodes (temporalité) spécifiques pour le contrôle du respect de l'interdiction du rejet d'hydrocarbures.",
        "Est-ce que {prep_q}, des procédures de contrôle sont spécifiquement décrites pour garantir ou assurer le respect de l'interdiction du rejet d'hydrocarbures ? Si ces procédures existent, vérifiez si elles précisent des lieux ou zones spécifiques pour le contrôle du respect de l'interdiction du rejet d'hydrocarbures.",
    ]
}

# ─────────────────────────────────────────────────────────────────────────────
# UTILITAIRES EXCEL ET SLUGIFY
# ─────────────────────────────────────────────────────────────────────────────

def _slugify(text: str) -> str:
    """Normalise et convertit une chaîne en un slug valide pour les noms de fichiers."""
    normalized = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('utf-8')
    return normalized.lower().replace(" ", "_").replace("'", "")

def _pays_sheet_name(pays_nom: str) -> str:
    """Nom de feuille Excel valide (max 31 chars, pas de caractères interdits)."""
    forbidden = r"\/*?:[]"
    name = pays_nom
    for ch in forbidden:
        name = name.replace(ch, "_")
    return name[:31]

def _clean_for_excel(val: str) -> str:
    """Nettoie une chaîne de caractères pour éviter les IllegalCharacterError d'openpyxl."""
    if not isinstance(val, str):
        return val
    val = re.sub(r'[\x01-\x08]', '• ', val)
    return re.sub(r'[\x00\x0b\x0c\x0e-\x1f]', '', val)

_HEADER_FONT       = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_FILL_Q     = PatternFill("solid", start_color="2F4F8F")   # bleu marine – Question
_HEADER_FILL_RAG_C = PatternFill("solid", start_color="4A7C59")   # vert sombre – RAG Classique
_HEADER_FILL_RAG_N = PatternFill("solid", start_color="7B3F6E")   # violet – RAG Neurosymbolique
_HEADER_ALIGN      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN        = Alignment(vertical="top", wrap_text=True)
_BORDER_THIN       = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_ALT_FILL = PatternFill("solid", start_color="F2F2F2")   # gris clair – lignes paires

def _style_sheet(ws, col_widths: List[int], header_fills: Optional[List[PatternFill]] = None):
    """Applique les styles sur la feuille active."""
    if header_fills is None:
        header_fills = [_HEADER_FILL_Q, _HEADER_FILL_RAG_C, _HEADER_FILL_RAG_N]

    for col_idx, (header, fill) in enumerate(zip(ws[1], header_fills), 1):
        header.font      = _HEADER_FONT
        header.fill      = fill
        header.alignment = _HEADER_ALIGN
        header.border    = _BORDER_THIN

    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for cell in row:
            cell.alignment = _CELL_ALIGN
            cell.border    = _BORDER_THIN
            if fill:
                cell.fill = fill

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 80


# ─────────────────────────────────────────────────────────────────────────────
# CLASSE BATCH QUERY SYSTEM
# ─────────────────────────────────────────────────────────────────────────────

class BatchQuerySystem:
    """Orchestre l'interrogation automatique du système RAG par batch."""

    def __init__(self):
        self.output_dir = Path("results/batch_queries")
        self.output_dir.mkdir(parents=True, exist_ok=True)

        existing_runs = []
        for f in self.output_dir.glob("*_run*.xlsx"):
            try:
                run_num = int(f.stem.split('_run')[-1])
                existing_runs.append(run_num)
            except ValueError:
                pass
        self.run_id = max(existing_runs) + 1 if existing_runs else 1

        self.run_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.questions_file = self.output_dir / "questions_generees.csv"
        self.results_file_classique = self.output_dir / f"resultats_audit_classique_run{self.run_id}_{self.run_timestamp}.csv"
        self.results_file_neuro     = self.output_dir / f"resultats_audit_neuro_run{self.run_id}_{self.run_timestamp}.csv"
        self.summary_file           = self.output_dir / f"resume_audit_run{self.run_id}_{self.run_timestamp}.json"

        self.fusion           = None
        self.llm_gen          = None
        self.query_analyzer   = None
        self.neo4j_bridge     = None
        self.graph_retriever  = None
        self.ontology_agent   = None
        self.dense_retriever  = None
        self.sparse_retriever = None

        logger.info("Initialisation du système de batch query")

    def init_rag_system(self) -> bool:
        """Initialise les composants du système RAG."""
        try:
            logger.info("Initialisation des composants RAG...")

            self.dense_retriever  = DenseRetriever()
            self.sparse_retriever = SparseRetriever()

            from rag.integration.neo4j_bridge import Neo4jBridge
            self.neo4j_bridge = Neo4jBridge.from_config()

            if not self.neo4j_bridge.is_ready():
                logger.warning("⚠ Neo4j Bridge non connecté. Le GraphRetriever sera ignoré.")
            else:
                logger.info("✓ Neo4j Bridge connecté")

            self.graph_retriever = Neo4jGraphRetriever(self.neo4j_bridge)
            self.ontology_agent  = Neo4jOntologyAgent(self.neo4j_bridge)
            self.fusion          = HybridFusion()
            logger.info("✓ Hybrid Fusion prêt")

            self.query_analyzer = QueryAnalyzer()
            logger.info("✓ Query Analyzer prêt")

            self.llm_gen = LLMGenerator()
            logger.info("✓ LLM Generator prêt")

            if not self.llm_gen.is_ollama_available():
                logger.warning("⚠ LLM non disponible (Ollama/Groq not reachable)")
                return False
            logger.info("✓ LLM disponible")

            return True
        except Exception as e:
            logger.error(f"❌ Erreur lors de l'initialisation: {e}", exc_info=True)
            return False

    def generer_questions(self) -> List[Dict]:
        """Génère les questions pour tous les pays avec des formulations naturelles."""
        questions = []

        print("\n📋 Génération des questions...")

        for pays_nom, pays_info in PAYS.items():
            prep   = pays_info["prep"]
            prep_q = pays_info["prep_q"]

            for interdiction, templates in INTERDICTIONS.items():
                for num_q, template in enumerate(templates, 1):
                    question = (
                        template
                        .replace("{prep}", prep)
                        .replace("{prep_q}", prep_q)
                    )
                    questions.append({
                        "Pays":            pays_nom,
                        "Code_Pays":       pays_info["code"],
                        "Interdiction":    interdiction,
                        "Numero_Question": num_q,
                        "Question":        question,
                    })

        with open(self.questions_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=questions[0].keys())
            writer.writeheader()
            writer.writerows(questions)

        print(f"✓ {len(questions)} questions générées et sauvegardées")
        return questions

    def interroger_question(self, question_data: Dict) -> Dict:
        """Interroge le système RAG (pipelines Classique et Neurosymbolique).

        Différence clé avec l'ancienne version :
          - RAG Classique : appel dense SANS expanded_terms (retrieval sémantique pur)
          - RAG Neuro     : appel dense AVEC expanded_terms + sparse + graph + ontologie
        Les deux pipelines sont maintenant totalement indépendants dès le retrieval.
        """
        question = question_data["Question"]

        try:
            if not self.query_analyzer or not self.fusion or not self.llm_gen:
                raise RuntimeError("Système RAG non initialisé")

            # ── 1. Analyse de la requête ──────────────────────────────────────
            # On extrait le pays purement à partir de la question (comme dans l'application réelle)
            intent, weights, country_filter, category_filter = self.query_analyzer.analyze(question)

            # ── 2. Construction des synonymes / expanded_terms (pour neuro uniquement) ──
            synonym_sets   = {}
            expanded_terms = []
            graph_entities = set()
            query_context  = None

            if self.ontology_agent and self.ontology_agent.is_ready():
                query_context = self.ontology_agent.prepare_query(question, intent=intent)
                synonym_sets = query_context.get("synonyms_map", {})
                graph_entities = query_context.get("graph_entities", set())
                expanded_terms = query_context.get("expanded_terms", [])

            _RETRIEVER_K = 12

            # ── 3. RAG CLASSIQUE — Dense PUR (sans expanded_terms) ────────────
            #
            #   On appelle le dense retriever directement sur la question brute,
            #   sans injecter aucun synonyme dans le vecteur de recherche.
            #   La fusion reçoit synonym_sets={} et enriched_context=None.
            #   Cela garantit un pipeline purement sémantique / vectoriel.

            dense_classic = []
            if self.dense_retriever and self.dense_retriever.is_ready():
                dense_classic = self.dense_retriever.retrieve(
                    question,
                    top_k=_RETRIEVER_K,
                    country_filter=country_filter,
                    # expanded_terms volontairement absent
                )

            # RAG CLASSIQUE STRICT: On ne passe PAS par fusion.fuse() pour avoir le retriever brut exact
            # On prend juste le top_k=6 du dense
            fusion_classic = dense_classic[:6]

            # Streaming pour éviter le post-processing strict qui casse le LLM
            response_classic_gen = self.llm_gen.generate(
                query=question,
                fusion_results=fusion_classic,
                intent=intent,
                enriched_context=None,
                stream=True,
            )
            if response_classic_gen.get("error"):
                answer_classic = "ERREUR"
            else:
                full_ans = "".join(list(response_classic_gen["stream_generator"]))
                answer_classic = re.sub(r'<analyse>.*?(?:</analyse>|$)', '', full_ans, flags=re.DOTALL).strip()

            # ── 4. RAG NEUROSYMBOLIQUE — Dense enrichi + Sparse + Graph + Ontologie ──
            #
            #   Le dense est appelé une seconde fois, cette fois AVEC expanded_terms.
            #   On ajoute le sparse et le graph retriever, la fusion reçoit les
            #   synonym_sets complets, l'ontology_boost et l'enriched_context.

            retriever_results_neuro = {}
            
            complex_intents = {
                "sanction_penale", "sanction_financiere", 
                "condition_temporelle", "condition_spatiale", 
                "controle_institution", "exploratory"
            }
            dense_expanded = expanded_terms if intent not in complex_intents else []

            if self.dense_retriever and self.dense_retriever.is_ready():
                retriever_results_neuro["dense"] = self.dense_retriever.retrieve(
                    question,
                    top_k=_RETRIEVER_K,
                    expanded_terms=dense_expanded,
                    country_filter=country_filter,
                )

            if self.sparse_retriever and self.sparse_retriever.is_ready():
                retriever_results_neuro["sparse"] = self.sparse_retriever.retrieve(
                    question,
                    top_k=_RETRIEVER_K,
                    expanded_terms=expanded_terms,
                    country_filter=country_filter,
                )

            # Le Graph Retriever est désactivé ici (comme dans Streamlit)
            # pour ne pas polluer la recherche, il ne sert qu'à l'enrichissement (OntologyAgent)

            fusion_neuro = self.fusion.fuse(
                retriever_results=retriever_results_neuro,
                weights=weights,
                graph_entities=graph_entities,
                synonym_sets=synonym_sets,
                ontology_boost=0.5,
                top_k=6,
                country_filter=country_filter,
                category_filter=category_filter,
                query=question,
                intent=intent,
                expanded_terms=expanded_terms,
            )

            enriched_context = None
            if self.ontology_agent and self.ontology_agent.is_ready():
                enriched_context = self.ontology_agent.enrich(question, fusion_neuro, query_context=query_context)

            # Streaming pour éviter le post-processing strict qui casse le LLM
            response_neuro_gen = self.llm_gen.generate(
                query=question,
                fusion_results=fusion_neuro,
                intent=intent,
                enriched_context=enriched_context,
                stream=True,
            )
            if response_neuro_gen.get("error"):
                answer_neuro = "ERREUR"
            else:
                full_ans = "".join(list(response_neuro_gen["stream_generator"]))
                answer_neuro = re.sub(r'<analyse>.*?(?:</analyse>|$)', '', full_ans, flags=re.DOTALL).strip()

            return {
                "Pays":                    question_data["Pays"],
                "Code_Pays":               question_data["Code_Pays"],
                "Interdiction":            question_data["Interdiction"],
                "Numero_Question":         question_data["Numero_Question"],
                "Question":                question,
                "Reponse_Classique":       answer_classic,
                "Reponse_Neurosymbolique": answer_neuro,
                "Modele":                  self.llm_gen.model_name,
                "Erreur":                  response_neuro_gen.get("error", "") or response_classic_gen.get("error", ""),
                "Timestamp":               datetime.now().isoformat(),
            }

        except Exception as e:
            logger.error(f"Erreur lors du traitement de la question: {e}", exc_info=True)
            return {
                "Pays":                    question_data["Pays"],
                "Code_Pays":               question_data["Code_Pays"],
                "Interdiction":            question_data["Interdiction"],
                "Numero_Question":         question_data["Numero_Question"],
                "Question":                question,
                "Reponse_Classique":       "ERREUR",
                "Reponse_Neurosymbolique": "ERREUR",
                "Modele":                  "N/A",
                "Erreur":                  str(e),
                "Timestamp":               datetime.now().isoformat(),
            }

    def sauvegarder_resultats(self, resultats: List[Dict]):
        """Sauvegarde les résultats dans deux CSV distincts."""
        if not resultats:
            logger.warning("Aucun résultat à sauvegarder")
            return

        res_classique = []
        res_neuro     = []
        for r in resultats:
            row_c = r.copy()
            row_c["Reponse"] = row_c.pop("Reponse_Classique")
            row_c.pop("Reponse_Neurosymbolique")
            res_classique.append(row_c)

            row_n = r.copy()
            row_n["Reponse"] = row_n.pop("Reponse_Neurosymbolique")
            row_n.pop("Reponse_Classique")
            res_neuro.append(row_n)

        with open(self.results_file_classique, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=res_classique[0].keys())
            writer.writeheader()
            writer.writerows(res_classique)

        with open(self.results_file_neuro, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=res_neuro[0].keys())
            writer.writeheader()
            writer.writerows(res_neuro)

    def mettre_a_jour_excel(self, resultats: List[Dict]):
        """Mise à jour incrémentale des fichiers Excel séparés par interdiction."""
        par_interdiction: Dict[str, List[Dict]] = {}
        for r in resultats:
            par_interdiction.setdefault(r["Interdiction"], []).append(r)

        for interdiction, rows in par_interdiction.items():
            slug       = _slugify(interdiction)
            excel_path = self.output_dir / f"{slug}_resultats_run{self.run_id}.xlsx"

            rows_by_pays: Dict[str, List[Dict]] = {}
            for r in rows:
                rows_by_pays.setdefault(r["Pays"], []).append(r)

            self.ecrire_excel(excel_path, rows_by_pays, mode="mixte")

    def ecrire_excel(self, output_path: Path, rows_by_pays: Dict[str, List[Dict]], mode: str = "mixte"):
        """Crée un fichier Excel avec une feuille par pays."""
        wb = Workbook()
        if wb.active is not None:
            wb.remove(wb.active)

        if mode == "classique":
            headers      = ["Question", "RAG Classique (dense pur)"]
            header_fills = [_HEADER_FILL_Q, _HEADER_FILL_RAG_C]
            col_widths   = [55, 100]
        elif mode == "neuro":
            headers      = ["Question", "RAG Neurosymbolique (dense + sparse + graph)"]
            header_fills = [_HEADER_FILL_Q, _HEADER_FILL_RAG_N]
            col_widths   = [55, 100]
        else:
            headers      = ["Question", "RAG Classique (dense pur)", "RAG Neurosymbolique (dense + sparse + graph)"]
            header_fills = [_HEADER_FILL_Q, _HEADER_FILL_RAG_C, _HEADER_FILL_RAG_N]
            col_widths   = [55, 60, 60]

        for pays_nom in PAYS.keys():
            rows = rows_by_pays.get(pays_nom, [])
            if not rows:
                continue

            ws = wb.create_sheet(title=_pays_sheet_name(pays_nom))
            ws.append(headers)
            ws.row_dimensions[1].height = 30

            for r in rows:
                if mode == "classique":
                    ws.append([_clean_for_excel(r["Question"]), _clean_for_excel(r["Reponse_Classique"])])
                elif mode == "neuro":
                    ws.append([_clean_for_excel(r["Question"]), _clean_for_excel(r["Reponse_Neurosymbolique"])])
                else:
                    ws.append([
                        _clean_for_excel(r["Question"]),
                        _clean_for_excel(r["Reponse_Classique"]),
                        _clean_for_excel(r["Reponse_Neurosymbolique"]),
                    ])

            _style_sheet(ws, col_widths=col_widths, header_fills=header_fills)

        wb.save(output_path)

    def generer_resume(self, resultats: List[Dict]):
        """Génère un résumé JSON des résultats."""
        resume = {
            "date_execution":       datetime.now().isoformat(),
            "total_questions":      len(resultats),
            "total_pays":           len(PAYS),
            "total_interdictions":  len(INTERDICTIONS),
            "erreurs_totales":      sum(
                1 for r in resultats
                if r["Reponse_Neurosymbolique"] == "ERREUR" or r["Reponse_Classique"] == "ERREUR"
            ),
            "reponses_oui_neuro":   sum(1 for r in resultats if "OUI" in str(r.get("Reponse_Neurosymbolique", "")).upper()),
            "reponses_non_neuro":   sum(1 for r in resultats if "NON" in str(r.get("Reponse_Neurosymbolique", "")).upper()),
            "reponses_oui_classic": sum(1 for r in resultats if "OUI" in str(r.get("Reponse_Classique", "")).upper()),
            "reponses_non_classic": sum(1 for r in resultats if "NON" in str(r.get("Reponse_Classique", "")).upper()),
            "par_pays":             {},
            "par_interdiction":     {},
        }

        for pays_nom in PAYS.keys():
            q_pays = [r for r in resultats if r["Pays"] == pays_nom]
            resume["par_pays"][pays_nom] = {
                "total":       len(q_pays),
                "erreurs":     sum(1 for r in q_pays if r["Reponse_Neurosymbolique"] == "ERREUR" or r["Reponse_Classique"] == "ERREUR"),
                "oui_neuro":   sum(1 for r in q_pays if "OUI" in str(r.get("Reponse_Neurosymbolique", "")).upper()),
                "non_neuro":   sum(1 for r in q_pays if "NON" in str(r.get("Reponse_Neurosymbolique", "")).upper()),
                "oui_classic": sum(1 for r in q_pays if "OUI" in str(r.get("Reponse_Classique", "")).upper()),
                "non_classic": sum(1 for r in q_pays if "NON" in str(r.get("Reponse_Classique", "")).upper()),
            }

        for interdiction in INTERDICTIONS.keys():
            q_i = [r for r in resultats if r["Interdiction"] == interdiction]
            resume["par_interdiction"][interdiction] = {
                "total":       len(q_i),
                "erreurs":     sum(1 for r in q_i if r["Reponse_Neurosymbolique"] == "ERREUR" or r["Reponse_Classique"] == "ERREUR"),
                "oui_neuro":   sum(1 for r in q_i if "OUI" in str(r.get("Reponse_Neurosymbolique", "")).upper()),
                "non_neuro":   sum(1 for r in q_i if "NON" in str(r.get("Reponse_Neurosymbolique", "")).upper()),
                "oui_classic": sum(1 for r in q_i if "OUI" in str(r.get("Reponse_Classique", "")).upper()),
                "non_classic": sum(1 for r in q_i if "NON" in str(r.get("Reponse_Classique", "")).upper()),
            }

        with open(self.summary_file, 'w', encoding='utf-8') as f:
            json.dump(resume, f, ensure_ascii=False, indent=2)

        logger.info(f"✓ Résumé sauvegardé dans: {self.summary_file}")
        return resume

    def afficher_resume(self, resume: Dict):
        """Affiche le résumé formaté dans la console."""
        print("\n" + "="*60)
        print("📊 RÉSUMÉ DE L'AUDIT")
        print("="*60)
        print(f"Total de questions : {resume['total_questions']}")
        print(f"Total de pays      : {resume['total_pays']}")
        print(f"Total interdictions: {resume['total_interdictions']}")
        print(f"Erreurs            : {resume['erreurs_totales']}")
        print(f"RAG Neuro  — OUI: {resume['reponses_oui_neuro']}  NON: {resume['reponses_non_neuro']}")
        print(f"RAG Class  — OUI: {resume['reponses_oui_classic']}  NON: {resume['reponses_non_classic']}")
        print("="*60 + "\n")

    def run(self, generate_only: bool = False, query_only: bool = False):
        """Exécute le pipeline complet."""
        try:
            if not query_only:
                questions = self.generer_questions()
            else:
                questions = []
                with open(self.questions_file, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        row["Numero_Question"] = int(row["Numero_Question"])
                        questions.append(row)
                print(f"✓ {len(questions)} questions chargées depuis {self.questions_file}")

            if generate_only:
                print("\n✅ Génération terminée. Utilisez --query-only pour interroger le système.")
                return

            if not self.init_rag_system():
                logger.error("❌ Impossible d'initialiser le système RAG")
                return

            resultats = []
            total = len(questions)
            print(f"\n🚀 Interrogation du système RAG ({total} questions)...\n")

            for idx, q_data in enumerate(questions, 1):
                print(
                    f"[{idx}/{total}] {q_data['Pays']} — {q_data['Interdiction']} "
                    f"(Q{q_data['Numero_Question']})...",
                    end=" ", flush=True,
                )
                result = self.interroger_question(q_data)
                resultats.append(result)
                print("✓")

                self.sauvegarder_resultats(resultats)
                self.mettre_a_jour_excel(resultats)

                if idx % 10 == 0:
                    logger.info(f"Progression: {idx}/{total} questions traitées")

            resume = self.generer_resume(resultats)
            self.afficher_resume(resume)
            print(f"✅ Fichiers de résultats créés dans: {self.output_dir}")

        except Exception as e:
            logger.error(f"❌ Erreur fatale: {e}", exc_info=True)
            raise


def main():
    parser = argparse.ArgumentParser(
        description="Interrogation automatique du système RAG par batch",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--generate-only", action="store_true", help="Générer les questions uniquement")
    parser.add_argument("--query-only",    action="store_true", help="Interroger les questions déjà générées")
    args = parser.parse_args()

    system = BatchQuerySystem()
    system.run(generate_only=args.generate_only, query_only=args.query_only)


if __name__ == "__main__":
    main()